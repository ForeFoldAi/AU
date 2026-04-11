#!/usr/bin/env python3
"""
BioTime Cloud Attendance & OT Report Generator  v2.0
Auinfocity | ZKTeco BioTime Cloud Integration
Department-specific rules: MEP / Security / House Keeping / Landscape / Pest Control
"""

import json
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from pathlib import Path
import calendar
import re
from collections import defaultdict
import sys
import argparse

# ══════════════════════════════════════════════════════════════════════════════
#  GLOBAL CONFIG
# ══════════════════════════════════════════════════════════════════════════════
CONFIG = {
    "base_url":     "https://auinfocity.itimedev.minervaiot.com",
    "company":      "auinfocity",
    "email":        "",
    "password":     "",
    "month":        11,
    "year":         2025,
    "company_name": "AUINFOCITY",
    "powered_by_line": "Powered by forefoldai.com",
    "site_name":    "CYBER TOWERS",
    "page_size":    100,
    "default_standard_hours": 8,
    "default_weekly_off_days": [6],
}

# ══════════════════════════════════════════════════════════════════════════════
#  DEPARTMENT RULES
#  Keys must exactly match department names returned by /personnel/api/departments/
#  If a department name doesn't match, default rules are applied and a warning is printed.
# ══════════════════════════════════════════════════════════════════════════════
# MEP / O&M: General 09:00–18:00, A 07:00–14:00, B 14:00–21:00, C 21:00–07:00
_MEP_ENGINE_RULES = {
    "standard_hours":                 9,
    "weekly_off_days":                [6],
    "weekly_off_adjacent":            True,
    "weekly_off_either_adjacent":     True,
    "mep_shift_engine":               True,
    "mep_general_sched_hours":        9.0,
    "mep_general_ot_min_extra":       3.0,
}

# Security: General 09:00–18:00, A4 08:00–20:00, C4 20:00–08:00 — no weekly off / PH
_SECURITY_ENGINE_RULES = {
    "standard_hours":            8,
    "weekly_off_days":           [],
    "weekly_off_adjacent":       False,
    "no_holidays":               True,
    "security_shift_engine":     True,
    "security_fixed_ot_hrs":     4.0,
    "security_std_duty_hrs":     8.0,
    "security_ot_min_session_hrs": 4.0,
}

# HK: General 08–17 / 09–18, A 06–15, B 12–21, C 21–06; W/O = either adjacent day duty
_HK_ENGINE_RULES = {
    "standard_hours":              9,
    "weekly_off_days":             [6],
    "weekly_off_adjacent":         True,
    "weekly_off_either_adjacent":  True,
    "hk_shift_engine":             True,
    "hk_a_continuation_ot_hrs":    6.0,
    "hk_bc_continuation_ot_hrs":   9.0,
}

# Landscape / Pest: General only 09:00–17:00 (8h); no OT; W/O = either adjacent day duty
_LP_ENGINE_RULES = {
    "standard_hours":              8,
    "weekly_off_days":             [6],
    "weekly_off_adjacent":         True,
    "weekly_off_either_adjacent":  True,
    "no_ot":                       True,
    "lp_shift_engine":             True,
}

DEPT_RULES = {

    # ── MEP / O&M (same rules; punches classified into G / A / B / C)
    "MEP": dict(_MEP_ENGINE_RULES),
    "O&M": dict(_MEP_ENGINE_RULES),
    "Operations & Maintenance": dict(_MEP_ENGINE_RULES),

    # ── Security ──────────────────────────────────────────────────────────────
    # G / A4 / C4 from punches; 4h fixed OT per A4 or C4 session only (not General).
    # Designation containing "Driver" → General-classified punches become A4 (C4 unchanged).
    "Security": dict(_SECURITY_ENGINE_RULES),

    # ── House Keeping ─────────────────────────────────────────────────────────
    # G/A/B/C from punches; no OT on General; A continuation +6h OT; B/C +9h (full day).
    "House Keeping": dict(_HK_ENGINE_RULES),

    # ── Landscape / Pest Control ──────────────────────────────────────────────
    "Landscape":    dict(_LP_ENGINE_RULES),
    "Pest Control": dict(_LP_ENGINE_RULES),
    # Common BioTime / template names → same rules (G only, no OT)
    "Nursary":      dict(_LP_ENGINE_RULES),
    "Nursery":      dict(_LP_ENGINE_RULES),
    "Gardners":     dict(_LP_ENGINE_RULES),
    "Gardeners":    dict(_LP_ENGINE_RULES),
}

# ── Attendance status codes ───────────────────────────────────────────────────
PRESENT  = "P"
PARTIAL  = "PP"
LEAVE    = "L"
WEEK_OFF = "W/O"
PUB_HOL  = "PH"


def _classify_mep_shift(login: datetime) -> str:
    """
    Classify a clock-in time into MEP shift (scheduled blocks):
    A 07:00–14:00, General 09:00–18:00, B 14:00–21:00, C 21:00–07:00.
    """
    m = login.hour * 60 + login.minute
    if m < 7 * 60:
        return "C"
    if m < 9 * 60:
        return "A"
    if m < 14 * 60:
        return "General"
    if m < 21 * 60:
        return "B"
    return "C"


def _mep_anchor_date(login: datetime, shift: str) -> date:
    """Overnight C ending before 07:00 is attributed to the shift start (previous day)."""
    if shift == "C" and login.hour < 7:
        return login.date() - timedelta(days=1)
    return login.date()


def _mep_shift_letter(shift_name: str) -> str:
    return "G" if shift_name == "General" else shift_name


def _mep_day_code_and_ot(day_sessions: list, gen_hrs: float, min_extra: float) -> tuple[str, int, float]:
    """
    Build display code (G, A, AC, ABC, G+OT) and OT day count / hours for one calendar day.
    """
    day_sessions = sorted(day_sessions, key=lambda s: s["login"])
    if not day_sessions:
        return LEAVE, 0, 0.0

    base_code = "".join(_mep_shift_letter(s["shift"]) for s in day_sessions)

    if len(day_sessions) == 1:
        s0 = day_sessions[0]
        if s0["shift"] == "General":
            extra = s0["hrs"] - gen_hrs
            if extra >= min_extra:
                return "G+OT", 1, round(extra, 2)
            return "G", 0, 0.0
        return base_code, 0, 0.0

    od = len(day_sessions) - 1
    ot_hrs = sum(s["hrs"] for s in day_sessions[1:])
    return base_code, od, round(ot_hrs, 2)


def _mep_status_worked_for_weekly_off(status) -> bool:
    if status in (LEAVE, WEEK_OFF, "", None):
        return False
    if status == PUB_HOL:
        return True
    return True


def _mep_status_counts_present_line(status) -> bool:
    if status in (LEAVE, WEEK_OFF, "", None):
        return False
    return status != PUB_HOL


def _mep_status_counts_man_day_row(status) -> bool:
    if status in (LEAVE, WEEK_OFF, "", None):
        return False
    return True


def _classify_security_shift(login: datetime) -> str:
    """
    Security clock-in → shift type:
    General 09:00–18:00, A4 08:00–20:00, C4 20:00–08:00 (next day).
    """
    m = login.hour * 60 + login.minute
    c4_evening_start = 19 * 60 + 30
    if m >= c4_evening_start or m < 8 * 60:
        return "C4"
    if m < 9 * 60:
        return "A4"
    return "General"


def _security_anchor_date(login: datetime, shift: str) -> date:
    if shift == "C4" and login.hour < 8:
        return login.date() - timedelta(days=1)
    return login.date()


def _security_display_token(shift_name: str) -> str:
    return "G" if shift_name == "General" else shift_name


def _designation_is_security_a4_driver(designation: str) -> bool:
    """Position/title contains 'driver' (e.g. Driver, Van Driver) → Security A4 roster."""
    return "driver" in (designation or "").strip().casefold()


def _security_day_code_and_ot(
    day_sessions: list,
    fixed_ot: float,
    min_session_hrs: float,
) -> tuple[str, int, float]:
    """Build G / A4 / C4 / A4C4 and OT (4h per qualifying A4 or C4 session only)."""
    day_sessions = sorted(day_sessions, key=lambda s: s["login"])
    if not day_sessions:
        return LEAVE, 0, 0.0

    base_code = "".join(_security_display_token(s["shift"]) for s in day_sessions)

    ot_sessions = [
        s
        for s in day_sessions
        if s["shift"] in ("A4", "C4") and s["hrs"] >= min_session_hrs
    ]
    if not ot_sessions:
        return base_code, 0, 0.0

    od = len(ot_sessions)
    oh = round(fixed_ot * od, 2)
    return base_code, od, oh


def _classify_hk_shift(login: datetime) -> str:
    """
    House Keeping clock-in:
    C 21:00–06:00, B 12:00–21:00, A 06:00–08:00, General 08:00–12:00 (08–17 / 09–18 starts).
    """
    m = login.hour * 60 + login.minute
    if m >= 21 * 60 or m < 6 * 60:
        return "C"
    if 12 * 60 <= m < 21 * 60:
        return "B"
    if 6 * 60 <= m < 8 * 60:
        return "A"
    if 8 * 60 <= m < 12 * 60:
        return "General"
    return "General"


def _hk_anchor_date(login: datetime, shift: str) -> date:
    if shift == "C" and login.hour < 6:
        return login.date() - timedelta(days=1)
    return login.date()


def _hk_display_token(shift_name: str) -> str:
    return "G" if shift_name == "General" else shift_name


def _hk_day_code_and_ot(
    day_sessions: list,
    a_ot: float,
    bc_ot: float,
) -> tuple[str, int, float]:
    """
    Codes: G, A, B, C, BC, … No OT on General; A in continuation → +a_ot hrs;
    B/C continuation (no General in pair) → +bc_ot full-day hours per step.
    """
    day_sessions = sorted(day_sessions, key=lambda s: s["login"])
    if not day_sessions:
        return LEAVE, 0, 0.0

    base_code = "".join(_hk_display_token(s["shift"]) for s in day_sessions)

    od = 0
    oh = 0.0
    for i in range(1, len(day_sessions)):
        prev = day_sessions[i - 1]["shift"]
        curr = day_sessions[i]["shift"]
        if prev == "General" or curr == "General":
            continue
        if prev == "A" or curr == "A":
            od += 1
            oh += a_ot
        elif prev in ("B", "C") or curr in ("B", "C"):
            od += 1
            oh += bc_ot

    return base_code, od, round(oh, 2)


# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "dark_blue":  "1F4E79",
    "med_blue":   "2E75B6",
    "light_blue": "BDD7EE",
    "present":    "C6EFCE",   # green
    "weekoff":    "FFEB9C",   # amber
    "leave":      "FFC7CE",   # red
    "partial":    "DDEBF7",   # pale blue
    "pubhol":     "E2EFDA",   # mint
    "ot_day":     "FFF2CC",   # yellow tint — Present day that also earned OT
    "ot_row_bg":  "FFFDE7",   # very pale yellow — OT hours sub-row background
    "summ_bg":    "D6DCE4",
    "total_bg":   "EBF3FB",
    "dept_hdr":   "2F5496",
    "sun_hdr":    "C00000",   # red header for weekly-off days
    "white":      "FFFFFF",
    "black":      "000000",
}


def _report_cell_bg_shift_code(code, od: int, fill_map: dict) -> str | None:
    """MEP / Security / HK letter codes (G, BC, A4C4, …) + OT highlighting."""
    if code in fill_map:
        base = fill_map[code]
    elif code and code not in (LEAVE, WEEK_OFF):
        base = C["present"]
    else:
        base = None
    if od and code not in (LEAVE, WEEK_OFF, "", None):
        return C["ot_day"]
    return base


# ══════════════════════════════════════════════════════════════════════════════
#  API CLIENT
# ══════════════════════════════════════════════════════════════════════════════
class BioTimeClient:
    def __init__(self, cfg):
        self.base   = cfg["base_url"].rstrip("/")
        self.cfg    = cfg
        self.token  = None
        self.scheme = "Token"

    def authenticate(self):
        email, pw, co = self.cfg["email"], self.cfg["password"], self.cfg.get("company", "")
        attempts = [
            (f"{self.base}/api-token-auth/",
             {"company": co, "email": email, "password": pw}, "Token"),
            (f"{self.base}/jwt-api-token-auth/",
             {"company": co, "email": email, "password": pw}, "JWT"),
            (f"{self.base}/staff-jwt-api-token-auth/",
             {"company": co, "username": email, "password": pw}, "JWT"),
            (f"{self.base}/staff-api-token-auth/",
             {"username": email, "password": pw}, "Token"),
            (f"{self.base}/api-token-auth/",
             {"email": email, "password": pw}, "Token"),
        ]
        last = ""
        for url, payload, scheme in attempts:
            try:
                r = requests.post(url, json=payload, timeout=30)
                if r.ok:
                    data = r.json()
                    tok = data.get("token") or data.get("access") or data.get("jwt")
                    if tok:
                        self.token, self.scheme = tok, scheme
                        print(f"OK  Auth: [{scheme}] via /{url.split('/')[-2]}/")
                        return
                else:
                    last = f"\n  {url} -> {r.status_code}: {r.text[:300]}"
            except requests.exceptions.RequestException as e:
                last = f"\n  {url} -> {e}"
        print("FAIL  Authentication failed:" + last)
        sys.exit(1)

    def _hdrs(self):
        return {"Authorization": f"{self.scheme} {self.token}",
                "Content-Type": "application/json"}

    def _get_all(self, url, params=None):
        params = dict(params or {})
        params.setdefault("page_size", self.cfg["page_size"])
        params["page"] = 1
        out = []
        while True:
            r = requests.get(url, headers=self._hdrs(), params=params, timeout=60)
            r.raise_for_status()
            d = r.json()
            out.extend(d.get("data") or d.get("results") or [])
            if not d.get("next"):
                break
            params["page"] += 1
        return out

    def employees(self):
        v = self._get_all(f"{self.base}/personnel/api/employees/")
        print(f"OK  Employees  : {len(v)}")
        return v

    def departments(self):
        v = self._get_all(f"{self.base}/personnel/api/departments/")
        print(f"OK  Departments: {len(v)}")
        return v

    def positions(self):
        try:
            v = self._get_all(f"{self.base}/personnel/api/positions/")
            print(f"OK  Positions  : {len(v)}")
            return v
        except Exception:
            return []

    def transactions(self, sd: date, ed: date):
        v = self._get_all(f"{self.base}/iclock/api/transactions/",
                          {"start_time": f"{sd} 00:00:00", "end_time": f"{ed} 23:59:59"})
        print(f"OK  Transactions: {len(v)}")
        return v


# ══════════════════════════════════════════════════════════════════════════════
#  SHIFT DETECTOR  — infers A / B / C / General from first punch-in hour
# ══════════════════════════════════════════════════════════════════════════════
class ShiftDetector:
    def __init__(self, shifts_cfg: dict):
        self.shifts = shifts_cfg   # {"General": {"punch_in_hours": (7,11)}, ...}

    def detect(self, first_punch: datetime) -> str:
        h = first_punch.hour
        for name, scfg in self.shifts.items():
            lo, hi = scfg.get("punch_in_hours", (0, 0))
            in_window = (lo <= h < hi) if lo < hi else (h >= lo or h < hi)
            if in_window:
                return name
        return "General"


# ══════════════════════════════════════════════════════════════════════════════
#  DEPARTMENT-AWARE ATTENDANCE CALCULATOR
# ══════════════════════════════════════════════════════════════════════════════
class AttendanceCalculator:

    def __init__(self, dept_name: str, rule: dict, designation: str = ""):
        self.dept         = dept_name
        self.rule         = rule
        self.designation  = designation or ""
        self.std_hrs      = rule.get("standard_hours", 8)
        self.off_days     = set(rule.get("weekly_off_days", [6]))
        self.no_ot        = rule.get("no_ot", False)
        self.detector     = ShiftDetector(rule.get("shifts", {}))

    @staticmethod
    def _parse(s: str):
        s = s.replace("T", " ")[:19]
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                pass
        return None

    # ── OT logic ─────────────────────────────────────────────────────────────
    def _ot(self, punches: list, shift: str):
        """
        Returns (ot_day_count:int, ot_hours:float) for one worked day.
        """
        if len(punches) < 2:
            return 0, 0.0

        work_hrs = (punches[-1] - punches[0]).total_seconds() / 3600
        extra    = max(0.0, work_hrs - self.std_hrs)

        # ── SECURITY (legacy path only if security_shift_engine is off) ──
        if self.dept == "Security" and not self.rule.get("security_shift_engine"):
            fixed = self.rule.get("daily_fixed_ot_hrs", 4)
            return (1, float(fixed)) if work_hrs >= self.std_hrs else (0, 0.0)

        # ── Global no-OT ─────────────────────────────────────────────────
        if self.no_ot:
            return 0, 0.0

        shifts_cfg = self.rule.get("shifts", {})
        shift_cfg  = shifts_cfg.get(shift, {})

        # ── House Keeping (legacy only without hk_shift_engine) ──────────
        if self.dept == "House Keeping" and not self.rule.get("hk_shift_engine"):
            if shift_cfg.get("no_ot"):
                return 0, 0.0
            threshold = shift_cfg.get("double_shift_threshold", 0)
            ot_hrs    = float(shift_cfg.get("double_shift_ot_hrs", 0))
            if threshold and work_hrs >= threshold:
                return 1, ot_hrs
            return 0, 0.0

        # ── Other departments: generic OT if extra hours exist ────────────
        return (1, round(extra, 2)) if extra > 0 else (0, 0.0)

    # ── Weekly-off adjacency rule ─────────────────────────────────────────────
    def _adjacent_rule(self, attendance: dict) -> dict:
        """
        Weekly-off validation against adjacent calendar days.

        Default: duty required on **both** previous and next day.
        MEP / O&M / House Keeping / Landscape / Pest (shift engine + either flag): OR adjacent day.
        """
        mep = self.rule.get("mep_shift_engine")
        hk = self.rule.get("hk_shift_engine")
        lp = self.rule.get("lp_shift_engine")
        either = self.rule.get("weekly_off_either_adjacent")

        def _worked(st) -> bool:
            if mep or hk or lp:
                return _mep_status_worked_for_weekly_off(st)
            return st in (PRESENT, PARTIAL)

        revised = dict(attendance)
        for d, status in attendance.items():
            if status != WEEK_OFF:
                continue
            prev_ok = _worked(attendance.get(d - timedelta(days=1)))
            next_ok = _worked(attendance.get(d + timedelta(days=1)))
            if either:
                if not (prev_ok or next_ok):
                    revised[d] = LEAVE
            else:
                if not (prev_ok and next_ok):
                    revised[d] = LEAVE
        return revised

    def _compute_mep(self, txns: list, month: int, year: int):
        """Pair punches into sessions, classify shifts, merge by anchor day."""
        gen_hrs = float(self.rule.get("mep_general_sched_hours", 9.0))
        min_extra = float(self.rule.get("mep_general_ot_min_extra", 3.0))

        pts = sorted(
            x for x in (self._parse(t.get("punch_time", "")) for t in txns) if x
        )
        num_days = calendar.monthrange(year, month)[1]
        month_start = date(year, month, 1)
        month_end = date(year, month, num_days)

        sessions: list[dict] = []
        for i in range(0, len(pts) - 1, 2):
            a, b = pts[i], pts[i + 1]
            if b <= a:
                continue
            sh = _classify_mep_shift(a)
            ad = _mep_anchor_date(a, sh)
            sessions.append(
                {
                    "login": a,
                    "logout": b,
                    "hrs": (b - a).total_seconds() / 3600.0,
                    "shift": sh,
                    "anchor": ad,
                }
            )

        by_anchor: dict = defaultdict(list)
        for s in sessions:
            if month_start <= s["anchor"] <= month_end:
                by_anchor[s["anchor"]].append(s)

        attendance: dict = {}
        ot_days: dict = {}
        ot_hours: dict = {}

        for day in range(1, num_days + 1):
            d = date(year, month, day)
            is_wo = d.weekday() in self.off_days
            day_sess = by_anchor.get(d, [])
            if day_sess:
                code, od, oh = _mep_day_code_and_ot(day_sess, gen_hrs, min_extra)
                attendance[d], ot_days[d], ot_hours[d] = code, od, oh
            elif is_wo:
                attendance[d], ot_days[d], ot_hours[d] = WEEK_OFF, 0, 0.0
            else:
                attendance[d], ot_days[d], ot_hours[d] = LEAVE, 0, 0.0

        if self.rule.get("weekly_off_adjacent"):
            attendance = self._adjacent_rule(attendance)

        return attendance, ot_days, ot_hours

    def _compute_security(self, txns: list, month: int, year: int):
        """Pair punches, classify G / A4 / C4, fixed 4h OT per A4/C4 session only."""
        fixed_ot = float(self.rule.get("security_fixed_ot_hrs", 4.0))
        min_sess = float(self.rule.get("security_ot_min_session_hrs", 4.0))

        pts = sorted(
            x for x in (self._parse(t.get("punch_time", "")) for t in txns) if x
        )
        num_days = calendar.monthrange(year, month)[1]
        month_start = date(year, month, 1)
        month_end = date(year, month, num_days)

        sessions: list[dict] = []
        for i in range(0, len(pts) - 1, 2):
            a, b = pts[i], pts[i + 1]
            if b <= a:
                continue
            sh = _classify_security_shift(a)
            if _designation_is_security_a4_driver(self.designation) and sh == "General":
                sh = "A4"
            ad = _security_anchor_date(a, sh)
            sessions.append(
                {
                    "login": a,
                    "logout": b,
                    "hrs": (b - a).total_seconds() / 3600.0,
                    "shift": sh,
                    "anchor": ad,
                }
            )

        by_anchor: dict = defaultdict(list)
        for s in sessions:
            if month_start <= s["anchor"] <= month_end:
                by_anchor[s["anchor"]].append(s)

        attendance: dict = {}
        ot_days: dict = {}
        ot_hours: dict = {}

        for day in range(1, num_days + 1):
            d = date(year, month, day)
            is_wo = d.weekday() in self.off_days
            day_sess = by_anchor.get(d, [])
            if day_sess:
                code, od, oh = _security_day_code_and_ot(day_sess, fixed_ot, min_sess)
                attendance[d], ot_days[d], ot_hours[d] = code, od, oh
            elif is_wo:
                attendance[d], ot_days[d], ot_hours[d] = WEEK_OFF, 0, 0.0
            else:
                attendance[d], ot_days[d], ot_hours[d] = LEAVE, 0, 0.0

        if self.rule.get("weekly_off_adjacent"):
            attendance = self._adjacent_rule(attendance)

        return attendance, ot_days, ot_hours

    def _compute_hk(self, txns: list, month: int, year: int):
        """House Keeping: pair punches, G/A/B/C, continuation OT (no OT on General)."""
        a_ot = float(self.rule.get("hk_a_continuation_ot_hrs", 6.0))
        bc_ot = float(self.rule.get("hk_bc_continuation_ot_hrs", 9.0))

        pts = sorted(
            x for x in (self._parse(t.get("punch_time", "")) for t in txns) if x
        )
        num_days = calendar.monthrange(year, month)[1]
        month_start = date(year, month, 1)
        month_end = date(year, month, num_days)

        sessions: list[dict] = []
        for i in range(0, len(pts) - 1, 2):
            a, b = pts[i], pts[i + 1]
            if b <= a:
                continue
            sh = _classify_hk_shift(a)
            ad = _hk_anchor_date(a, sh)
            sessions.append(
                {
                    "login": a,
                    "logout": b,
                    "hrs": (b - a).total_seconds() / 3600.0,
                    "shift": sh,
                    "anchor": ad,
                }
            )

        by_anchor: dict = defaultdict(list)
        for s in sessions:
            if month_start <= s["anchor"] <= month_end:
                by_anchor[s["anchor"]].append(s)

        attendance: dict = {}
        ot_days: dict = {}
        ot_hours: dict = {}

        for day in range(1, num_days + 1):
            d = date(year, month, day)
            is_wo = d.weekday() in self.off_days
            day_sess = by_anchor.get(d, [])
            if day_sess:
                code, od, oh = _hk_day_code_and_ot(day_sess, a_ot, bc_ot)
                attendance[d], ot_days[d], ot_hours[d] = code, od, oh
            elif is_wo:
                attendance[d], ot_days[d], ot_hours[d] = WEEK_OFF, 0, 0.0
            else:
                attendance[d], ot_days[d], ot_hours[d] = LEAVE, 0, 0.0

        if self.rule.get("weekly_off_adjacent"):
            attendance = self._adjacent_rule(attendance)

        return attendance, ot_days, ot_hours

    def _compute_lp(self, txns: list, month: int, year: int):
        """Landscape / Pest Control: single General shift (09:00–17:00); always G, no OT."""
        pts = sorted(
            x for x in (self._parse(t.get("punch_time", "")) for t in txns) if x
        )
        num_days = calendar.monthrange(year, month)[1]
        month_start = date(year, month, 1)
        month_end = date(year, month, num_days)

        sessions: list[dict] = []
        for i in range(0, len(pts) - 1, 2):
            a, b = pts[i], pts[i + 1]
            if b <= a:
                continue
            sessions.append(
                {
                    "login": a,
                    "logout": b,
                    "hrs": (b - a).total_seconds() / 3600.0,
                    "shift": "General",
                    "anchor": a.date(),
                }
            )

        by_anchor: dict = defaultdict(list)
        for s in sessions:
            if month_start <= s["anchor"] <= month_end:
                by_anchor[s["anchor"]].append(s)

        attendance: dict = {}
        ot_days: dict = {}
        ot_hours: dict = {}

        for day in range(1, num_days + 1):
            d = date(year, month, day)
            is_wo = d.weekday() in self.off_days
            if by_anchor.get(d):
                attendance[d], ot_days[d], ot_hours[d] = "G", 0, 0.0
            elif is_wo:
                attendance[d], ot_days[d], ot_hours[d] = WEEK_OFF, 0, 0.0
            else:
                attendance[d], ot_days[d], ot_hours[d] = LEAVE, 0, 0.0

        if self.rule.get("weekly_off_adjacent"):
            attendance = self._adjacent_rule(attendance)

        return attendance, ot_days, ot_hours

    # ── Main compute ──────────────────────────────────────────────────────────
    def compute(self, txns: list, month: int, year: int):
        if self.rule.get("mep_shift_engine"):
            return self._compute_mep(txns, month, year)
        if self.rule.get("security_shift_engine"):
            return self._compute_security(txns, month, year)
        if self.rule.get("hk_shift_engine"):
            return self._compute_hk(txns, month, year)
        if self.rule.get("lp_shift_engine"):
            return self._compute_lp(txns, month, year)

        by_date: dict = defaultdict(list)
        for t in txns:
            pt = self._parse(t.get("punch_time", ""))
            if pt:
                by_date[pt.date()].append(pt)

        num_days = calendar.monthrange(year, month)[1]
        attendance, ot_days, ot_hours = {}, {}, {}

        for day in range(1, num_days + 1):
            d       = date(year, month, day)
            punches = sorted(by_date.get(d, []))
            is_wo   = d.weekday() in self.off_days

            if punches:
                shift = self.detector.detect(punches[0])
                if len(punches) >= 2:
                    wh     = (punches[-1] - punches[0]).total_seconds() / 3600
                    status = PRESENT if wh >= self.std_hrs else PARTIAL
                else:
                    wh, status = 0.0, PARTIAL
                od, oh = self._ot(punches, shift)
            elif is_wo:
                status, od, oh = WEEK_OFF, 0, 0.0
            else:
                status, od, oh = LEAVE, 0, 0.0

            attendance[d], ot_days[d], ot_hours[d] = status, od, oh

        if self.rule.get("weekly_off_adjacent"):
            attendance = self._adjacent_rule(attendance)

        return attendance, ot_days, ot_hours

    # ── Summarize ─────────────────────────────────────────────────────────────
    def summarize(self, attendance: dict, ot_days: dict, ot_hours: dict) -> dict:
        if (
            self.rule.get("mep_shift_engine")
            or self.rule.get("security_shift_engine")
            or self.rule.get("hk_shift_engine")
            or self.rule.get("lp_shift_engine")
        ):
            present = sum(
                1 for v in attendance.values() if _mep_status_counts_present_line(v)
            )
            pp = 0
            wo = sum(1 for v in attendance.values() if v == WEEK_OFF)
            ph = sum(1 for v in attendance.values() if v == PUB_HOL)
            total_ot = sum(ot_days.values())
            return {
                "present":        present,
                "wo":             wo,
                "ot":             total_ot,
                "ph":             ph,
                "pp":             pp,
                "total_present":  present + wo + ph,
                "total_man_days": present + pp + total_ot,
            }

        cnt = lambda code: sum(1 for v in attendance.values() if v == code)
        present = cnt(PRESENT)
        pp = cnt(PARTIAL)
        wo = cnt(WEEK_OFF)
        ph = cnt(PUB_HOL)
        total_ot = sum(ot_days.values())
        return {
            "present":        present,
            "wo":             wo,
            "ot":             total_ot,
            "ph":             ph,
            "pp":             pp,
            "total_present":  present + wo + ph,
            "total_man_days": present + pp + total_ot,
        }


# ══════════════════════════════════════════════════════════════════════════════
#  DEPARTMENT → EXCEL SHEET TITLE (BioTime names may differ from rule keys)
# ══════════════════════════════════════════════════════════════════════════════
def sheet_base_name_for_department(dept_name: str) -> str:
    """
    Map API department label to the Excel worksheet name (the tab at the bottom
    of the workbook) and to the subtitle row inside that sheet.
    O&M-style names → MEP; landscape / gardening → Nursary (template spelling).
    """
    raw = (dept_name or "").strip()
    if not raw:
        return "Department"
    cf = raw.casefold()
    compact_amp = cf.replace(" ", "")

    if (
        re.search(r"\bo\s*&\s*m\b", cf)
        or re.search(r"\bo\s+and\s+m\b", cf)
        or "o&m" in compact_amp
        or ("operations" in cf and "maintenance" in cf)
    ):
        return "MEP"

    if any(
        k in cf
        for k in (
            "landscape",
            "nursary",
            "nursery",
            "gardener",
            "gardening",
            "garden maintenance",
        )
    ):
        return "Nursary"

    return raw


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT GENERATOR
# ══════════════════════════════════════════════════════════════════════════════
class ReportGenerator:
    DAY_ABBR = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]

    def __init__(self, cfg):
        self.cfg      = cfg
        self.month    = cfg["month"]
        self.year     = cfg["year"]
        self.num_days = calendar.monthrange(self.year, self.month)[1]
        self.mo_name  = calendar.month_name[self.month].upper()
        _t = Side(style="thin")
        self.BDR = Border(left=_t, right=_t, top=_t, bottom=_t)

    def _fill(self, h):
        return PatternFill("solid", start_color=h)

    def _font(self, bold=False, sz=8, fg=C["black"]):
        return Font(name="Arial", bold=bold, size=sz, color=fg)

    def _align(self, h="center", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    def _c(self, ws, row, col, val="", bold=False, sz=8,
           fg=C["black"], bg=None, ha="center", wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = self._font(bold, sz, fg)
        c.alignment = self._align(ha, wrap)
        c.border = self.BDR
        if bg:
            c.fill = self._fill(bg)
        return c

    def _mc(self, ws, r1, c1, r2, c2, val="", bold=False, sz=8,
            fg=C["black"], bg=None, ha="center", wrap=False):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        return self._c(ws, r1, c1, val, bold, sz, fg, bg, ha, wrap)

    @staticmethod
    def _excel_sheet_title(raw: str, used: set[str]) -> str:
        """Excel sheet name: max 31 chars, no []:*?/\\"""
        t = "".join("_" if c in "[]:*?/\\" else c for c in (raw or "").strip()) or "Department"
        if len(t) > 31:
            t = t[:31]
        base = t
        n = 1
        while t in used:
            suffix = f" ({n})"
            t = (base[: max(0, 31 - len(suffix))] + suffix).strip() or f"Dept_{n}"
            n += 1
        used.add(t)
        return t

    def _apply_sheet_layout(self, ws, dept_name: str, total_cols: int, fixed: int):
        """Header rows, column widths, freeze panes (one department per sheet)."""
        ws.freeze_panes = "E5"
        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 5
        for i in range(self.num_days):
            ws.column_dimensions[get_column_letter(fixed + 1 + i)].width = 3.8
        for i, w in enumerate([7, 4, 5, 4, 8, 10]):
            ws.column_dimensions[get_column_letter(fixed + self.num_days + 1 + i)].width = w
        for r in [1, 2, 3, 4]:
            ws.row_dimensions[r].height = {1: 24, 2: 18, 3: 16, 4: 14}[r]

        company = str(self.cfg.get("company_name") or "")
        powered = str(self.cfg.get("powered_by_line") or "Powered by forefoldai.com")
        if total_cols <= 1:
            self._mc(
                ws, 1, 1, 1, total_cols, company,
                bold=True, sz=14, fg=C["white"], bg=C["dark_blue"], ha="left",
            )
        else:
            mid = total_cols // 2
            self._mc(
                ws, 1, 1, 1, mid, company,
                bold=True, sz=14, fg=C["white"], bg=C["dark_blue"], ha="left",
            )
            self._mc(
                ws, 1, mid + 1, 1, total_cols, powered,
                bold=True, sz=11, fg=C["white"], bg=C["dark_blue"], ha="right",
            )
        self._mc(ws, 2, 1, 2, total_cols,
                 f"{self.cfg['site_name']}  —  {dept_name}  —  "
                 f"ATTENDANCE FOR THE MONTH OF {self.mo_name} {self.year}",
                 bold=True, sz=10, fg=C["white"], bg=C["med_blue"])

        for col, hdr in enumerate(["S.No", "Employee Name", "Designation", "W/D"], 1):
            self._mc(ws, 3, col, 4, col, hdr,
                     bold=True, sz=8, fg=C["white"], bg=C["dark_blue"], wrap=True)

        for i in range(self.num_days):
            d = date(self.year, self.month, i + 1)
            col = fixed + 1 + i
            hc = C["dark_blue"]
            self._c(ws, 3, col, i + 1, bold=True, sz=8, fg=C["white"], bg=hc)
            self._c(ws, 4, col, self.DAY_ABBR[d.weekday()],
                    bold=True, sz=7, fg=C["white"], bg=hc)

        for i, lbl in enumerate(
                ["Present", "W/O", "OT", "PH", "Total\nPresent", "Total Man\ndays"]):
            self._mc(ws, 3, fixed + self.num_days + 1 + i,
                     4, fixed + self.num_days + 1 + i,
                     lbl, bold=True, sz=8, fg=C["white"], bg=C["dark_blue"], wrap=True)

    @staticmethod
    def _sort_rows_by_designation_then_name(rows: list) -> list:
        """Ascending A–Z by designation; blank designation last; then by name."""

        def key(e):
            des = str(e.get("designation") or "").strip()
            name = str(e.get("name") or "").strip()
            return (0 if des else 1, des.casefold(), name.casefold())

        return sorted(rows, key=key)

    def _write_employee_block(self, ws, dept_rows: list, fixed: int, total_cols: int) -> int:
        """Write designation band rows + employee rows; returns next row after last employee."""
        fill_map = {PRESENT: C["present"], WEEK_OFF: C["weekoff"],
                    LEAVE: C["leave"], PARTIAL: C["partial"], PUB_HOL: C["pubhol"]}
        rule0 = _get_rule(dept_rows[0]["dept"]) if dept_rows else {}
        uses_shift_codes = bool(
            rule0.get("mep_shift_engine")
            or rule0.get("security_shift_engine")
            or rule0.get("hk_shift_engine")
            or rule0.get("lp_shift_engine")
        )
        lp_g_only = bool(rule0.get("lp_shift_engine"))
        ordered = self._sort_rows_by_designation_then_name(dept_rows)
        current_row = 5
        sno = 1
        prev_group: object | str = object()

        for emp in ordered:
            des = str(emp.get("designation") or "").strip()
            group_label = des if des else "—"
            if group_label != prev_group:
                rband = current_row
                ws.row_dimensions[rband].height = 13
                self._mc(
                    ws, rband, 1, rband, total_cols,
                    f"  {group_label}",
                    bold=True, sz=8, fg=C["white"], bg=C["dept_hdr"], ha="left",
                )
                current_row += 1
                prev_group = group_label

            r = current_row
            r2 = current_row + 1
            ws.row_dimensions[r].height = 14
            ws.row_dimensions[r2].height = 11

            self._mc(ws, r, 1, r2, 1, sno, sz=8)
            self._mc(ws, r, 2, r2, 2, emp["name"], sz=8, ha="left")
            self._mc(ws, r, 3, r2, 3, emp.get("designation", ""), sz=7, ha="left")
            self._mc(ws, r, 4, r2, 4, emp.get("wo_label") or "—", sz=7)

            for i in range(self.num_days):
                col = fixed + 1 + i
                d = date(self.year, self.month, i + 1)
                code = emp["attendance"].get(d, "")
                od = emp["ot_days"].get(d, 0)
                if lp_g_only:
                    if code not in (LEAVE, WEEK_OFF, PUB_HOL, "", None):
                        code = "G"
                    od = 0
                    bg = fill_map.get(code, C["present"])
                elif uses_shift_codes:
                    bg = _report_cell_bg_shift_code(code, od, fill_map)
                else:
                    bg = C["ot_day"] if (od and code == PRESENT) else fill_map.get(code)
                self._c(ws, r, col, code, sz=7, bg=bg)

            for i in range(self.num_days):
                col = fixed + 1 + i
                oh = emp["ot_hours"].get(date(self.year, self.month, i + 1), 0.0)
                if lp_g_only:
                    self._c(ws, r2, col, "", sz=7, bg=C["ot_row_bg"])
                elif oh:
                    self._c(ws, r2, col, round(oh, 1), sz=7, fg=C["black"], bg=C["ot_day"])
                else:
                    self._c(ws, r2, col, "", sz=7, bg=C["ot_row_bg"])

            summ = dict(emp["summary"])
            if lp_g_only:
                summ["ot"] = 0
                summ["total_man_days"] = summ["present"] + summ["pp"]
            keys = ["present", "wo", "ot", "ph", "total_present", "total_man_days"]
            for i, key in enumerate(keys):
                self._c(ws, r, fixed + self.num_days + 1 + i,
                        summ[key], bold=True, sz=8, bg=C["total_bg"])

            total_ot_hrs = 0.0 if lp_g_only else round(sum(emp["ot_hours"].values()), 1)
            for i, key in enumerate(keys):
                if key == "ot":
                    self._c(ws, r2, fixed + self.num_days + 1 + i,
                            total_ot_hrs if total_ot_hrs else "",
                            bold=True, sz=7,
                            bg=C["ot_day"] if total_ot_hrs else C["ot_row_bg"])
                else:
                    self._c(ws, r2, fixed + self.num_days + 1 + i,
                            "", sz=7, bg=C["ot_row_bg"])

            current_row += 2
            sno += 1
        return current_row

    def _write_dept_summary_rows(self, ws, dept_rows: list, start_row: int, fixed: int) -> int:
        """Daily and column totals for employees on this sheet only."""
        summ_keys = ["present", "wo", "ot", "ph", "total_present", "total_man_days"]
        summ_labels = ["Present", "W/O", "OT", "PH", "Total Present", "Total Man days"]
        current_row = start_row
        rule0 = _get_rule(dept_rows[0]["dept"]) if dept_rows else {}
        uses_shift_codes = bool(
            rule0.get("mep_shift_engine")
            or rule0.get("security_shift_engine")
            or rule0.get("hk_shift_engine")
            or rule0.get("lp_shift_engine")
        )

        for label, key in zip(summ_labels, summ_keys):
            r = current_row
            ws.row_dimensions[r].height = 13
            self._mc(ws, r, 1, r, fixed, label,
                     bold=True, sz=8, bg=C["summ_bg"], ha="right")

            for i in range(self.num_days):
                d = date(self.year, self.month, i + 1)
                if key == "present":
                    if uses_shift_codes:
                        val = sum(
                            1 for e in dept_rows
                            if _mep_status_counts_present_line(e["attendance"].get(d))
                        )
                    else:
                        val = sum(1 for e in dept_rows if e["attendance"].get(d) == PRESENT)
                elif key == "wo":
                    val = sum(1 for e in dept_rows if e["attendance"].get(d) == WEEK_OFF)
                elif key == "ot":
                    val = sum(e["ot_days"].get(d, 0) for e in dept_rows)
                elif key == "ph":
                    val = sum(1 for e in dept_rows if e["attendance"].get(d) == PUB_HOL)
                elif key == "total_present":
                    if uses_shift_codes:
                        val = sum(
                            1 for e in dept_rows
                            if _mep_status_counts_present_line(e["attendance"].get(d))
                            or e["attendance"].get(d) == WEEK_OFF
                            or e["attendance"].get(d) == PUB_HOL
                        )
                    else:
                        val = sum(
                            1 for e in dept_rows
                            if e["attendance"].get(d) in (PRESENT, WEEK_OFF, PUB_HOL)
                        )
                else:
                    if uses_shift_codes:
                        val = sum(
                            1 for e in dept_rows
                            if _mep_status_counts_man_day_row(e["attendance"].get(d))
                        )
                    else:
                        val = sum(
                            1 for e in dept_rows
                            if e["attendance"].get(d) in (PRESENT, PARTIAL)
                        )
                self._c(ws, r, fixed + 1 + i,
                        val if val else "", bold=True, sz=7, bg=C["summ_bg"])

            for i, k in enumerate(summ_keys):
                col2 = fixed + self.num_days + 1 + i
                if k == key:
                    self._c(ws, r, col2, sum(e["summary"][key] for e in dept_rows),
                            bold=True, sz=8, bg=C["light_blue"])
                else:
                    c2 = ws.cell(row=r, column=col2)
                    c2.border = self.BDR
                    c2.fill = self._fill(C["summ_bg"])
            current_row += 1
        return current_row

    def generate(self, emp_rows: list, output: str):
        FIXED = 4
        SUMM = 6
        TOTAL = FIXED + self.num_days + SUMM

        user_weekoffs = load_user_weekoffs()
        rows_for_report = []
        for e in emp_rows:
            rule = _get_rule(e["dept"])
            calc = AttendanceCalculator(
                e["dept"], rule, str(e.get("designation") or "")
            )
            att = e["attendance"]
            if rule.get("weekly_off_adjacent"):
                att = calc._adjacent_rule(att)
            summ = calc.summarize(att, e["ot_days"], e["ot_hours"])
            ec = e.get("emp_code", "")
            wo_label = wo_label_for_employee(user_weekoffs, str(ec), rule.get("weekly_off_days", []))
            rows_for_report.append({**e, "attendance": att, "summary": summ, "wo_label": wo_label})

        by_dept = defaultdict(list)
        for e in rows_for_report:
            by_dept[e["dept"]].append(e)

        wb = openpyxl.Workbook()
        if not by_dept:
            ws = wb.active
            ws.title = "Attendance"
            ws["A1"] = "No employees in this report."
            wb.save(output)
            print(f"OK  Report saved -> {output}")
            return

        wb.remove(wb.active)
        used_titles: set[str] = set()
        for dept_name in sorted(by_dept.keys()):
            dept_rows = by_dept[dept_name]
            # Worksheet tab name (Excel UI) — sanitized + unique in _excel_sheet_title
            worksheet_name = sheet_base_name_for_department(dept_name)
            tab_title = self._excel_sheet_title(worksheet_name, used_titles)
            ws = wb.create_sheet(title=tab_title)
            ws.title = tab_title  # same as create_sheet title; makes “tab = worksheet name” explicit
            self._apply_sheet_layout(ws, worksheet_name, TOTAL, FIXED)
            next_row = self._write_employee_block(ws, dept_rows, FIXED, TOTAL)
            self._write_dept_summary_rows(ws, dept_rows, next_row, FIXED)

        wb.save(output)
        print(f"OK  Report saved -> {output}  ({len(by_dept)} department sheet(s))")


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def _resolve(emp, key, fkeys, lkp):
    val = emp.get(key)
    if isinstance(val, dict):
        for fk in fkeys:
            if val.get(fk):
                return val[fk]
    entry = lkp.get(str(val)) if val is not None else None
    if entry:
        for fk in fkeys:
            if entry.get(fk):
                return entry[fk]
    for fk in fkeys:
        if emp.get(fk):
            return emp[fk]
    return ""

def _name(emp):
    fn = (emp.get("first_name") or "").strip()
    ln = (emp.get("last_name")  or "").strip()
    return (f"{fn} {ln}".strip() or
            emp.get("emp_name", emp.get("name", f"EMP-{emp.get('emp_code', '?')}")))


def employee_area(emp: dict) -> str:
    """Resolve work area from a BioTime employee payload (same rules as the GUI Employee model)."""
    area_list = emp.get("area") or []
    if area_list and isinstance(area_list, list) and len(area_list) > 0:
        a0 = area_list[0]
        if isinstance(a0, dict):
            return (a0.get("area_name") or a0.get("name") or "General") or "General"
        if isinstance(a0, str) and a0.strip():
            return a0.strip()
    v = emp.get("area_name")
    if isinstance(v, str) and v.strip():
        return v.strip()
    return "General"


def _wo_label(off_days):
    abbr = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
    return "/".join(abbr[d] for d in off_days) if off_days else "—"


# Same path as forefold_attendance_gui.weekoff.store — Weekly Off tab saves here.
USER_WEEKOFFS_PATH = Path.home() / ".forefold" / "weekoffs.json"

_WO_DAY_NAMES = [
    "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday",
]
_WO_DAY_ABBR = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]


def load_user_weekoffs(path: Path | None = None) -> dict[str, str | None]:
    """Load per-employee weekly-off day names from the GUI Weekly Off tab store."""
    p = path or USER_WEEKOFFS_PATH
    if not p.exists():
        return {}
    try:
        with p.open(encoding="utf-8") as f:
            data = json.load(f)
        raw = data.get("weekoffs") or {}
        # JSON keys are strings; normalize to str and strip for consistent lookup.
        return {str(k).strip(): v for k, v in raw.items()}
    except Exception:
        return {}


def _lookup_saved_weekoff_day(
    user_weekoffs: dict[str, str | None],
    emp_code: str,
) -> str | None:
    """
    Resolve the saved weekday for an employee. Keys in weekoffs.json must match
    personnel emp_code/id, but APIs may differ on padding (e.g. '042' vs '42').
    """
    if not user_weekoffs or not str(emp_code).strip():
        return None
    ec = str(emp_code).strip()

    def _val(k: str) -> str | None:
        v = user_weekoffs.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
        return None

    if (d := _val(ec)) is not None:
        return d

    # Numeric id variants (leading zeros, int string forms)
    if ec.isdigit():
        n = int(ec)
        for k, v in user_weekoffs.items():
            ks = str(k).strip()
            if not ks.isdigit():
                continue
            try:
                if int(ks) == n and isinstance(v, str) and v.strip():
                    return v.strip()
            except ValueError:
                continue

    return None


def _day_name_to_wd_abbr(day_name: str) -> str | None:
    try:
        return _WO_DAY_ABBR[_WO_DAY_NAMES.index(day_name)]
    except ValueError:
        return None


def wo_label_for_employee(
    user_weekoffs: dict[str, str | None],
    emp_code: str,
    dept_off_days: list[int],
) -> str:
    """
    W/D column: use only the day saved in the Weekly Off tab.
    If not set/invalid, display em dash (no department fallback).
    """
    saved = _lookup_saved_weekoff_day(user_weekoffs or {}, emp_code)
    if saved:
        abbr = _day_name_to_wd_abbr(saved)
        if abbr:
            return abbr
    return "—"


def employee_weekly_off_days(
    user_weekoffs: dict[str, str | None],
    emp_code: str,
) -> list[int]:
    """Return weekly-off weekday index list for one employee from saved tab data."""
    saved = _lookup_saved_weekoff_day(user_weekoffs or {}, emp_code)
    if not saved:
        return []
    try:
        return [_WO_DAY_NAMES.index(saved)]
    except ValueError:
        return []


def _department_is_landscape_pest_family(dept_name: str) -> bool:
    """Match gardener / nursery / landscape / pest labels from BioTime or templates."""
    dn = (dept_name or "").strip().casefold()
    if not dn:
        return False
    if "pest control" in dn or dn == "pest":
        return True
    if "landscape" in dn:
        return True
    for w in ("nursary", "nursery", "gardener", "gardeners", "gardening"):
        if w in dn:
            return True
    return False


def _get_rule(dept_name):
    if dept_name in DEPT_RULES:
        return DEPT_RULES[dept_name]
    dn = dept_name.lower()
    if "operations" in dn and "maintenance" in dn:
        return dict(_MEP_ENGINE_RULES)
    if _department_is_landscape_pest_family(dept_name):
        return dict(_LP_ENGINE_RULES)
    for k, v in DEPT_RULES.items():
        if k.lower() in dn or dn in k.lower():
            return v
    return {"standard_hours": CONFIG["default_standard_hours"],
            "weekly_off_days": CONFIG["default_weekly_off_days"],
            "weekly_off_adjacent": False}


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    p = argparse.ArgumentParser(description="AU Infocity - Vendor Attendance & OT Report v2")
    p.add_argument("--month",    type=int, default=CONFIG["month"])
    p.add_argument("--year",     type=int, default=CONFIG["year"])
    p.add_argument("--email",    default=CONFIG["email"])
    p.add_argument("--password", default=CONFIG["password"])
    p.add_argument("--company",  default=CONFIG["company"])
    p.add_argument("--output",   default=None)
    args = p.parse_args()

    cfg = {**CONFIG, "month": args.month, "year": args.year,
           "email": args.email, "password": args.password, "company": args.company}
    mo_name = calendar.month_name[cfg["month"]]
    output  = args.output or f"Attendance_{mo_name}_{cfg['year']}.xlsx"

    # Auth + fetch
    client = BioTimeClient(cfg)
    client.authenticate()
    raw_emps  = client.employees()
    raw_depts = client.departments()
    raw_pos   = client.positions()

    depts_map = {str(d["id"]): d for d in raw_depts}
    pos_map   = {str(p["id"]): p for p in raw_pos}

    num_days = calendar.monthrange(cfg["year"], cfg["month"])[1]
    sd, ed   = date(cfg["year"], cfg["month"], 1), date(cfg["year"], cfg["month"], num_days)
    raw_txns = client.transactions(sd, ed)

    txns_by_emp = defaultdict(list)
    for t in raw_txns:
        ec = str(t.get("emp_code", ""))
        if ec:
            txns_by_emp[ec].append(t)

    # Build per-employee rows
    emp_rows    = []
    unknown_depts = set()
    user_weekoffs = load_user_weekoffs()

    for emp in raw_emps:
        ec        = str(emp.get("emp_code", emp.get("id", "")))
        dept_name = _resolve(emp, "department", ["dept_name", "name"], depts_map) or "General"
        desig     = _resolve(emp, "position",   ["position_name", "name"], pos_map)
        rule      = _get_rule(dept_name)

        if dept_name not in DEPT_RULES:
            unknown_depts.add(dept_name)

        weekly_off_days = employee_weekly_off_days(user_weekoffs, ec)
        effective_rule = {**rule, "weekly_off_days": weekly_off_days}
        calc = AttendanceCalculator(dept_name, effective_rule, desig)
        att, od, oh = calc.compute(txns_by_emp.get(ec, []), cfg["month"], cfg["year"])
        summ        = calc.summarize(att, od, oh)

        emp_rows.append({
            "emp_code":    ec,
            "name":        _name(emp),
            "designation": desig,
            "dept":        dept_name,
            "wo_label":    wo_label_for_employee(user_weekoffs, ec, weekly_off_days),
            "attendance":  att,
            "ot_days":     od,
            "ot_hours":    oh,
            "summary":     summ,
        })

    if unknown_depts:
        print(f"\nWARN  Departments not in DEPT_RULES (default rules used): "
              f"{', '.join(sorted(unknown_depts))}")
        print("      Add them to DEPT_RULES in this script to apply specific rules.\n")

    # Console summary
    print(f"\n{'─'*58}")
    print(f"  {mo_name} {cfg['year']}  |  {len(emp_rows)} employees")
    print(f"  {'Department':<22}  {'Present':>7}  {'OT':>5}  {'ManDays':>8}")
    print(f"  {'─'*22}  {'─'*7}  {'─'*5}  {'─'*8}")
    for dn in sorted(set(e["dept"] for e in emp_rows)):
        g = [e for e in emp_rows if e["dept"] == dn]
        print(f"  {dn:<22}  {sum(e['summary']['present'] for e in g):>7}"
              f"  {sum(e['summary']['ot'] for e in g):>5}"
              f"  {sum(e['summary']['total_man_days'] for e in g):>8}")
    print(f"  {'─'*22}  {'─'*7}  {'─'*5}  {'─'*8}")
    print(f"  {'TOTAL':<22}  {sum(e['summary']['present'] for e in emp_rows):>7}"
          f"  {sum(e['summary']['ot'] for e in emp_rows):>5}"
          f"  {sum(e['summary']['total_man_days'] for e in emp_rows):>8}")
    print(f"{'─'*58}\n")

    ReportGenerator(cfg).generate(emp_rows, output)
    print(f"Done  ->  {output}")


if __name__ == "__main__":
    main()
