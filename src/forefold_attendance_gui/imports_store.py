"""Import storage: parse three Excel masters, persist rows, resolve shift timing."""

from __future__ import annotations

import json
import re
import shutil
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

KIND_SHIFT_TIMES = "shift_times"
KIND_TIMETABLE = "timetable"
KIND_EMPLOYEE_SCHEDULE = "employee_schedule"

_ALLOWED_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}

_STORE_DIR = Path.home() / ".forefold" / "imports"
_MANIFEST = _STORE_DIR / "manifest.json"

_JSON_FILES = {
    KIND_TIMETABLE: _STORE_DIR / "timetable_rows.json",
    KIND_SHIFT_TIMES: _STORE_DIR / "shift_rows.json",
    KIND_EMPLOYEE_SCHEDULE: _STORE_DIR / "schedule_rows.json",
}


def kind_label(kind: str) -> str:
    return {
        KIND_SHIFT_TIMES: "Shift master (Shift Name, Timetable, …)",
        KIND_TIMETABLE: "Shift time table (Name, Check In, Check Out, …)",
        KIND_EMPLOYEE_SCHEDULE: "Employee scheduled shifts",
    }.get(kind, kind)


def load_manifest() -> dict:
    if not _MANIFEST.exists():
        return {}
    try:
        return json.loads(_MANIFEST.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_manifest(data: dict) -> None:
    _STORE_DIR.mkdir(parents=True, exist_ok=True)
    _MANIFEST.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def _norm_header(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _norm_key(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _header_col_index(headers: list[Any], *aliases: str) -> int:
    """Return 1-based column index for first header matching any alias."""
    norm_aliases = [_norm_header(a) for a in aliases]
    for col_idx, raw in enumerate(headers, start=1):
        h = _norm_header(raw)
        if not h:
            continue
        for a in norm_aliases:
            if h == a or a in h or h in a:
                return col_idx
    raise ValueError(
        f"Missing column (expected one of: {', '.join(aliases)}). Found headers: {headers!r}"
    )


def _optional_col_index(headers: list[Any], *aliases: str) -> int | None:
    try:
        return _header_col_index(headers, *aliases)
    except ValueError:
        return None


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    if isinstance(val, time):
        return val.strftime("%H:%M")
    if isinstance(val, datetime):
        if val.hour or val.minute or val.second:
            return val.strftime("%H:%M")
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    return str(val).strip()


def _row_headers(ws, row_idx: int) -> list[Any]:
    """One sheet row trimmed of trailing empty cells (for use as column headers)."""
    if row_idx < 1 or row_idx > (ws.max_row or 0):
        return []
    row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
    headers = list(row)
    while headers and headers[-1] is None:
        headers.pop()
    return headers


def _find_data_header_row(
    ws,
    row_ok: Callable[[list[Any]], bool],
    *,
    max_scan_rows: int = 40,
    sheet_hint: str = "This sheet",
) -> tuple[int, list[Any]]:
    """Scan from row 1 for the first row that looks like a real header (title rows skipped)."""
    last_nonempty: list[Any] = []
    limit = min(max(ws.max_row or 1, 1), max_scan_rows)
    for r in range(1, limit + 1):
        headers = _row_headers(ws, r)
        if not headers or not any(_norm_header(h) for h in headers):
            continue
        last_nonempty = headers
        if row_ok(headers):
            return r, headers
    shown = last_nonempty if last_nonempty else "(no non-empty row in scan range)"
    raise ValueError(
        f"{sheet_hint}: could not find a data header row in the first {max_scan_rows} rows. "
        'If row 1 is only a title (e.g. "Employee Schedule"), put column names on the next row. '
        f"Last non-empty row scanned: {shown!r}"
    )


def _timetable_header_row_ok(headers: list[Any]) -> bool:
    try:
        _header_col_index(
            headers,
            "name (shift name)",
            "name(shift name)",
            "name",
            "shift name",
        )
        _header_col_index(headers, "check in", "checkin", "check-in")
        _header_col_index(headers, "check out", "checkout", "check-out")
        return True
    except ValueError:
        return False


def _shift_header_row_ok(headers: list[Any]) -> bool:
    try:
        _header_col_index(headers, "shift name", "shiftname")
        _header_col_index(headers, "timetable")
        return True
    except ValueError:
        return False


def _schedule_header_row_ok(headers: list[Any]) -> bool:
    try:
        _header_col_index(headers, "employee id", "employeeid", "emp id", "emp code")
        _header_col_index(headers, "shift name", "shiftname")
        _header_col_index(headers, "start date", "startdate")
        _header_col_index(headers, "end date", "enddate")
        return True
    except ValueError:
        return False


def parse_timetable_workbook(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        header_row, headers = _find_data_header_row(
            ws, _timetable_header_row_ok, sheet_hint="Timetable"
        )
        c_name = _header_col_index(
            headers,
            "name (shift name)",
            "name(shift name)",
            "name",
            "shift name",
        )
        c_type = _optional_col_index(headers, "type")
        c_in = _header_col_index(headers, "check in", "checkin", "check-in")
        c_out = _header_col_index(headers, "check out", "checkout", "check-out")
        c_work = _optional_col_index(headers, "work time", "worktime")
        c_break = _optional_col_index(headers, "break time", "breaktime")
        c_day = _optional_col_index(headers, "workday", "work day")
        c_wtype = _optional_col_index(headers, "work type", "worktype")
        c_fh = _optional_col_index(
            headers, "first half (check out time)", "first half", "first half check out"
        )
        c_sh = _optional_col_index(
            headers, "second half (check in time)", "second half", "second half check in"
        )

        rows: list[dict[str, str]] = []
        for r in range(header_row + 1, ws.max_row + 1):
            name = _cell_str(ws.cell(r, c_name).value)
            if not name:
                continue
            row = {
                "name": name,
                "type": _cell_str(ws.cell(r, c_type).value) if c_type else "",
                "check_in": _cell_str(ws.cell(r, c_in).value),
                "check_out": _cell_str(ws.cell(r, c_out).value),
                "work_time": _cell_str(ws.cell(r, c_work).value) if c_work else "",
                "break_time": _cell_str(ws.cell(r, c_break).value) if c_break else "",
                "work_day": _cell_str(ws.cell(r, c_day).value) if c_day else "",
                "work_type": _cell_str(ws.cell(r, c_wtype).value) if c_wtype else "",
                "first_half_check_out": _cell_str(ws.cell(r, c_fh).value) if c_fh else "",
                "second_half_check_in": _cell_str(ws.cell(r, c_sh).value) if c_sh else "",
            }
            rows.append(row)
        return rows
    finally:
        wb.close()


def parse_shift_workbook(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        header_row, headers = _find_data_header_row(
            ws, _shift_header_row_ok, sheet_hint="Shift master"
        )
        c_sn = _header_col_index(headers, "shift name", "shiftname")
        c_tt = _header_col_index(headers, "timetable")
        c_unit = _optional_col_index(headers, "unit")
        c_cycle = _optional_col_index(headers, "cycle")
        c_auto = _optional_col_index(headers, "auto shift", "autoshift")

        rows: list[dict[str, str]] = []
        for r in range(header_row + 1, ws.max_row + 1):
            sn = _cell_str(ws.cell(r, c_sn).value)
            if not sn:
                continue
            rows.append(
                {
                    "shift_name": sn,
                    "timetable": _cell_str(ws.cell(r, c_tt).value),
                    "unit": _cell_str(ws.cell(r, c_unit).value) if c_unit else "",
                    "cycle": _cell_str(ws.cell(r, c_cycle).value) if c_cycle else "",
                    "auto_shift": _cell_str(ws.cell(r, c_auto).value) if c_auto else "",
                }
            )
        return rows
    finally:
        wb.close()


def parse_schedule_workbook(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        header_row, headers = _find_data_header_row(
            ws, _schedule_header_row_ok, sheet_hint="Employee schedule"
        )
        c_eid = _header_col_index(headers, "employee id", "employeeid", "emp id", "emp code")
        c_fn = _optional_col_index(headers, "first name", "firstname")
        c_sn = _header_col_index(headers, "shift name", "shiftname")
        c_sd = _header_col_index(headers, "start date", "startdate")
        c_ed = _header_col_index(headers, "end date", "enddate")

        rows: list[dict[str, str]] = []
        for r in range(header_row + 1, ws.max_row + 1):
            eid = _cell_str(ws.cell(r, c_eid).value)
            if not eid:
                continue
            rows.append(
                {
                    "employee_id": eid,
                    "first_name": _cell_str(ws.cell(r, c_fn).value) if c_fn else "",
                    "shift_name": _cell_str(ws.cell(r, c_sn).value),
                    "start_date": _cell_str(ws.cell(r, c_sd).value),
                    "end_date": _cell_str(ws.cell(r, c_ed).value),
                }
            )
        return rows
    finally:
        wb.close()


def _write_json_rows(kind: str, rows: list[dict]) -> None:
    path = _JSON_FILES[kind]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(rows, indent=2, ensure_ascii=False), encoding="utf-8")


def load_timetable_rows() -> list[dict[str, str]]:
    p = _JSON_FILES[KIND_TIMETABLE]
    if not p.exists():
        return []
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return []


def load_shift_rows() -> list[dict[str, str]]:
    p = _JSON_FILES[KIND_SHIFT_TIMES]
    if not p.exists():
        return []
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return []


def load_schedule_rows() -> list[dict[str, str]]:
    p = _JSON_FILES[KIND_EMPLOYEE_SCHEDULE]
    if not p.exists():
        return []
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return []


def _timetable_by_name(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    for r in rows:
        k = _norm_key(r.get("name", ""))
        if k:
            out[k] = r
    return out


def _shift_by_shift_name(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    for r in rows:
        k = _norm_key(r.get("shift_name", ""))
        if k:
            out[k] = r
    return out


def _parse_date_loose(s: str) -> date | None:
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")[:10]).date()
    except ValueError:
        return None


def _schedule_by_employee(rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    by_e: dict[str, list[dict[str, str]]] = {}
    for r in rows:
        eid = _norm_key(r.get("employee_id", ""))
        if not eid:
            continue
        by_e.setdefault(eid, []).append(r)
    return by_e


def _match_emp_id(a: str, b: str) -> bool:
    na, nb = _norm_key(a), _norm_key(b)
    if na == nb:
        return True
    if na.isdigit() and nb.isdigit() and int(na) == int(nb):
        return True
    return False


def _pick_schedule_row(rows: list[dict[str, str]]) -> dict[str, str] | None:
    if not rows:
        return None
    today = date.today()
    in_range: list[dict[str, str]] = []
    for r in rows:
        sd = _parse_date_loose(r.get("start_date", ""))
        ed = _parse_date_loose(r.get("end_date", ""))
        if sd and ed and sd <= today <= ed:
            in_range.append(r)
    if in_range:
        return in_range[0]
    # Prefer latest end date
    best = None
    best_ed: date | None = None
    for r in rows:
        ed = _parse_date_loose(r.get("end_date", ""))
        if ed and (best_ed is None or ed > best_ed):
            best_ed = ed
            best = r
    return best or rows[0]


def format_shift_timing(shift_name: str, shifts: dict[str, dict], timetables: dict[str, dict]) -> str:
    if not shift_name.strip():
        return "—"
    srow = shifts.get(_norm_key(shift_name))
    if not srow:
        return "—"
    tt_name = (srow.get("timetable") or "").strip()
    if not tt_name:
        return "—"
    trow = timetables.get(_norm_key(tt_name))
    if not trow:
        return "—"
    ci, co = trow.get("check_in", ""), trow.get("check_out", "")
    if ci and co:
        return f"{ci} – {co}"
    if ci:
        return ci
    if co:
        return co
    wt = (trow.get("work_time") or "").strip()
    return wt or "—"


def enrich_employees_from_imports(employees: list[Any]) -> list[Any]:
    """Attach import_shift_name and shift_timing on Employee instances (dataclass replace)."""
    from dataclasses import replace

    from forefold_attendance_gui.dashboard.employees.model import Employee

    tt_rows = load_timetable_rows()
    sh_rows = load_shift_rows()
    sc_rows = load_schedule_rows()
    if not sc_rows:
        return [
            replace(
                e,
                import_shift_name="",
                shift_timing="—",
            )
            if isinstance(e, Employee)
            else e
            for e in employees
        ]

    timetables = _timetable_by_name(tt_rows)
    shifts = _shift_by_shift_name(sh_rows)
    by_emp = _schedule_by_employee(sc_rows)

    out: list[Any] = []
    for e in employees:
        if not isinstance(e, Employee):
            out.append(e)
            continue
        candidates: list[dict[str, str]] = []
        for k, lst in by_emp.items():
            if _match_emp_id(e.emp_id, k):
                candidates.extend(lst)
        row = _pick_schedule_row(candidates)
        if not row:
            out.append(
                replace(
                    e,
                    import_shift_name="",
                    shift_timing="—",
                )
            )
            continue
        sn = row.get("shift_name", "").strip()
        timing = format_shift_timing(sn, shifts, timetables) if sn else "—"
        out.append(
            replace(
                e,
                import_shift_name=sn,
                shift_timing=timing,
            )
        )
    return out


def import_one(kind: str, source_path: str | Path) -> dict:
    src = Path(source_path).expanduser().resolve()
    if src.suffix.lower() not in _ALLOWED_SUFFIXES:
        raise ValueError(
            f"Unsupported file type for {src.name}. Use .xlsx/.xlsm/.xltx/.xltm"
        )
    if not src.exists() or not src.is_file():
        raise ValueError(f"File not found: {src}")

    if kind == KIND_TIMETABLE:
        rows = parse_timetable_workbook(src)
    elif kind == KIND_SHIFT_TIMES:
        rows = parse_shift_workbook(src)
    elif kind == KIND_EMPLOYEE_SCHEDULE:
        rows = parse_schedule_workbook(src)
    else:
        raise ValueError(f"Unknown import kind: {kind}")

    _STORE_DIR.mkdir(parents=True, exist_ok=True)
    dest = _STORE_DIR / f"{kind}{src.suffix.lower()}"
    shutil.copy2(src, dest)
    _write_json_rows(kind, rows)

    wb = load_workbook(dest, read_only=True, data_only=True)
    try:
        sheet_name = wb.active.title
    finally:
        wb.close()

    return {
        "kind": kind,
        "source_path": str(src),
        "stored_path": str(dest),
        "sheet_name": sheet_name,
        "row_count": len(rows),
        "imported_at": _utc_now_iso(),
    }


def _utc_now_iso() -> str:
    from datetime import timezone

    return datetime.now(timezone.utc).isoformat()


def import_all(file_map: dict[str, str | Path]) -> dict[str, dict]:
    required = {KIND_SHIFT_TIMES, KIND_TIMETABLE, KIND_EMPLOYEE_SCHEDULE}
    missing = [k for k in sorted(required) if not file_map.get(k)]
    if missing:
        missing_lbl = ", ".join(kind_label(k) for k in missing)
        raise ValueError(f"Missing required file(s): {missing_lbl}")

    results = {}
    for kind in sorted(required):
        results[kind] = import_one(kind, file_map[kind])
    for r in results.values():
        r["imported_at"] = _utc_now_iso()

    manifest = load_manifest()
    manifest.update(results)
    _save_manifest(manifest)
    return results
